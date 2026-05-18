"""
FileParserService
=================
Parse file content → plain text để dùng cho chunking và OCR pipeline.

Supported file types (khớp với be_core ALLOWED_MIME_TYPES và FE accept):
  - PDF         (.pdf)  → text extraction (fitz) OR OCR fallback (Tesseract)
  - Word        (.docx) → python-docx
                (.doc)  ← KHÔNG hỗ trợ (old binary format, python-docx không đọc được)
  - Excel       (.xlsx) → openpyxl
                (.xls)  → xlrd (old binary format — openpyxl không hỗ trợ)
  - Text/MD     (.txt, .md, .markdown) → plain read + chardet encoding

OCR detection (PDF):
  - Nếu avg_chars/trang < OCR_TEXT_THRESHOLD → coi là scanned → Tesseract fallback
  - OCR_TEXT_THRESHOLD = 50 (mặc định)
"""
from __future__ import annotations

import asyncio
import os
from typing import Optional
from urllib.parse import urlparse

import httpx
import structlog

from app.core.exceptions import DocumentParsingException

log = structlog.get_logger(__name__)

# Nếu text extraction trung bình < N chars/trang → coi là scanned PDF
OCR_TEXT_THRESHOLD = 50

# File types được phép upload — PHẢI khớp với AllowedFileType trong schemas/ingest.py
# Lưu ý: .doc (old binary Word) KHÔNG được hỗ trợ, chỉ .docx
ALLOWED_FILE_TYPES = {"pdf", "docx", "xlsx", "xls", "txt", "md", "markdown"}


class FileParserService:

    @staticmethod
    def _is_remote_path(file_path: str) -> bool:
        parsed = urlparse(file_path)
        return parsed.scheme in {"http", "https"} and bool(parsed.netloc)

    @classmethod
    async def _download_remote_file(cls, file_url: str) -> bytes:
        timeout = httpx.Timeout(connect=10.0, read=60.0, write=10.0, pool=5.0)
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            response = await client.get(file_url)
            if response.status_code >= 400:
                raise DocumentParsingException(
                    f"Không tải được file từ URL: {file_url} (status={response.status_code})"
                )
            return response.content

    @classmethod
    async def parse(cls, file_path: str, file_type: str) -> str:
        """
        Parse a file and return its text content.

        Hỗ trợ: pdf, doc, docx, xlsx, xls, txt, md
        PDF scan tự động fallback sang Tesseract OCR.
        """
        if cls._is_remote_path(file_path):
            data = await cls._download_remote_file(file_path)
            parsed = urlparse(file_path)
            path_name = parsed.path.rsplit("/", 1)[-1] if parsed.path else ""
            guessed_name = path_name or f"remote_file.{file_type.lower().lstrip('.')}"
            return await cls.parse_bytes(data=data, file_name=guessed_name, mime_type="")

        if not os.path.exists(file_path):
            raise DocumentParsingException(f"File not found: {file_path}")

        file_type = file_type.lower().lstrip(".")

        if file_type not in ALLOWED_FILE_TYPES:
            raise DocumentParsingException(
                f"File type '{file_type}' không được hỗ trợ. "
                f"Chỉ chấp nhận: {', '.join(sorted(ALLOWED_FILE_TYPES))}"
            )

        try:
            if file_type == "pdf":
                return await cls._parse_pdf(file_path)
            elif file_type == "docx":
                return await cls._parse_docx(file_path)
            elif file_type == "xlsx":
                return await cls._parse_excel(file_path, use_xlrd=False)
            elif file_type == "xls":
                return await cls._parse_excel(file_path, use_xlrd=True)
            elif file_type in ("txt", "md", "markdown"):
                return await cls._parse_text(file_path)
            else:
                raise DocumentParsingException(f"Unsupported file type: {file_type}")
        except DocumentParsingException:
            raise
        except Exception as exc:
            raise DocumentParsingException(
                f"Failed to parse {file_type} file: {exc}"
            ) from exc

    @classmethod
    async def parse_bytes(cls, data: bytes, file_name: str, mime_type: str) -> str:
        """
        Parse từ raw bytes (dùng khi download từ ImageKit).
        Tự detect file type từ mime_type hoặc file_name extension.
        """
        import tempfile

        ext = _ext_from_mime(mime_type) or _ext_from_name(file_name)
        if not ext:
            raise DocumentParsingException(f"Không thể xác định file type: {file_name} ({mime_type})")

        # Ghi ra temp file rồi parse
        with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
            tmp.write(data)
            tmp_path = tmp.name

        try:
            return await cls.parse(tmp_path, ext)
        finally:
            os.unlink(tmp_path)

    # ── PDF ──────────────────────────────────────────────────────────────────

    @classmethod
    async def _parse_pdf(cls, file_path: str) -> str:
        """
        Đọc PDF bằng PyMuPDF.
        Nếu text quá ít (scanned PDF) → fallback sang Tesseract OCR.
        """
        import fitz

        text_parts: list[str] = []
        scanned_pages: list[int] = []

        doc = fitz.open(file_path)

        for page_num, page in enumerate(doc, start=1):
            page_text = page.get_text("text").strip()

            if len(page_text) < OCR_TEXT_THRESHOLD:
                # Trang này có vẻ là scan → đánh dấu để OCR sau
                scanned_pages.append(page_num)
                text_parts.append(f"[Page {page_num}][SCANNED]")
            else:
                text_parts.append(f"[Page {page_num}]\n{page_text}")

        doc.close()

        # Nếu > 50% trang là scan → dùng Tesseract cho toàn bộ file
        total_pages = len(text_parts)
        if total_pages > 0 and len(scanned_pages) / total_pages > 0.5:
            log.info(
                "file_parser.pdf_ocr_fallback",
                path=file_path,
                scanned_pages=len(scanned_pages),
                total_pages=total_pages,
            )
            return await cls._ocr_pdf(file_path)

        # Nếu chỉ 1 số trang bị scan, OCR từng trang đó
        if scanned_pages:
            log.info(
                "file_parser.pdf_partial_ocr",
                path=file_path,
                scanned_pages=scanned_pages,
            )
            ocr_texts = await cls._ocr_pdf_pages(file_path, scanned_pages)
            # Replace [SCANNED] placeholder bằng OCR text
            for i, page_num in enumerate(scanned_pages):
                text_parts[page_num - 1] = f"[Page {page_num}]\n{ocr_texts[i]}"

        full_text = "\n\n".join(p for p in text_parts if "[SCANNED]" not in p)
        log.debug("file_parser.pdf_parsed", path=file_path, pages=total_pages)
        return full_text

    @classmethod
    async def _ocr_pdf(cls, file_path: str) -> str:
        """OCR toàn bộ PDF bằng Tesseract. Dùng khi phần lớn trang là scan."""
        return await asyncio.to_thread(cls._ocr_pdf_sync, file_path)

    @staticmethod
    def _ocr_pdf_sync(file_path: str) -> str:
        """Synchronous OCR — chạy trong thread pool."""
        try:
            import pytesseract
            from pdf2image import convert_from_path
        except ImportError as e:
            raise DocumentParsingException(
                f"OCR dependencies chưa được cài: {e}. "
                "Cần: pytesseract, pdf2image, tesseract-ocr"
            )

        # Convert PDF → list of PIL images
        images = convert_from_path(file_path, dpi=200)

        text_parts: list[str] = []
        for i, img in enumerate(images, start=1):
            # Tesseract với ngôn ngữ tiếng Việt + tiếng Anh
            ocr_text = pytesseract.image_to_string(
                img,
                lang="vie+eng",
                config="--psm 3",  # Automatic page segmentation
            ).strip()

            if ocr_text:
                text_parts.append(f"[Page {i}]\n{ocr_text}")

        return "\n\n".join(text_parts)

    @classmethod
    async def _ocr_pdf_pages(cls, file_path: str, page_numbers: list[int]) -> list[str]:
        """OCR một số trang cụ thể trong PDF."""
        return await asyncio.to_thread(
            cls._ocr_specific_pages_sync, file_path, page_numbers
        )

    @staticmethod
    def _ocr_specific_pages_sync(file_path: str, page_numbers: list[int]) -> list[str]:
        try:
            import pytesseract
            from pdf2image import convert_from_path
        except ImportError as e:
            raise DocumentParsingException(f"OCR dependencies chưa được cài: {e}")

        results: list[str] = []
        for page_num in page_numbers:
            images = convert_from_path(
                file_path, dpi=200,
                first_page=page_num,
                last_page=page_num,
            )
            if images:
                text = pytesseract.image_to_string(
                    images[0], lang="vie+eng", config="--psm 3"
                ).strip()
                results.append(text)
            else:
                results.append("")
        return results

    # ── DOCX ─────────────────────────────────────────────────────────────────

    @classmethod
    async def _parse_docx(cls, file_path: str) -> str:
        """Parse DOCX dùng python-docx, lấy cả paragraphs và tables."""
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph

        doc = Document(file_path)
        parts: list[str] = []

        for block in doc.element.body:
            tag = block.tag.split("}")[-1] if "}" in block.tag else block.tag

            if tag == "p":
                para = Paragraph(block, doc)
                text = para.text.strip()
                if text:
                    parts.append(text)

            elif tag == "tbl":
                table = Table(block, doc)
                rows = []
                for row in table.rows:
                    row_text = " | ".join(cell.text.strip() for cell in row.cells)
                    if row_text.strip(" |"):
                        rows.append(row_text)
                if rows:
                    parts.append("\n".join(rows))

        full_text = "\n\n".join(parts)
        log.debug("file_parser.docx_parsed", path=file_path, blocks=len(parts))
        return full_text

    # ── Excel ─────────────────────────────────────────────────────────────────

    @classmethod
    async def _parse_excel(cls, file_path: str, use_xlrd: bool = False) -> str:
        """
        Parse Excel sang text dạng bảng.
        - .xlsx → openpyxl (use_xlrd=False)
        - .xls  → xlrd (use_xlrd=True) — openpyxl không đọc được old binary format
        """
        return await asyncio.to_thread(cls._parse_excel_sync, file_path, use_xlrd)

    @staticmethod
    def _parse_excel_sync(file_path: str, use_xlrd: bool = False) -> str:
        parts: list[str] = []

        if use_xlrd:
            # Old binary .xls format
            try:
                import xlrd
            except ImportError:
                raise DocumentParsingException(
                    "xlrd chưa được cài. Cần: pip install xlrd"
                )
            wb = xlrd.open_workbook(file_path)
            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                rows_text: list[str] = []
                for row_idx in range(ws.nrows):
                    cells = [
                        str(ws.cell_value(row_idx, col_idx)).strip()
                        for col_idx in range(ws.ncols)
                    ]
                    if any(cells):
                        rows_text.append(" | ".join(cells))
                if rows_text:
                    parts.append(f"[Sheet: {ws.name}]\n" + "\n".join(rows_text))
        else:
            # Modern .xlsx format
            try:
                import openpyxl
            except ImportError:
                raise DocumentParsingException(
                    "openpyxl chưa được cài. Cần: pip install openpyxl"
                )
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_text = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell).strip() if cell is not None else "" for cell in row]
                    if any(cells):
                        rows_text.append(" | ".join(cells))
                if rows_text:
                    parts.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text))
            wb.close()

        full_text = "\n\n".join(parts)
        log.debug("file_parser.excel_parsed", path=file_path, sheets=len(parts), xlrd=use_xlrd)
        return full_text

    # ── Text / Markdown ───────────────────────────────────────────────────────

    @classmethod
    async def _parse_text(cls, file_path: str) -> str:
        """Parse plain text với auto encoding detection."""
        import chardet

        with open(file_path, "rb") as f:
            raw = f.read()

        detected = chardet.detect(raw)
        encoding = detected.get("encoding") or "utf-8"

        try:
            text = raw.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            text = raw.decode("utf-8", errors="replace")

        ext = os.path.splitext(file_path)[-1].lower()
        if ext in (".md", ".markdown"):
            text = _flatten_markdown_tables(text)

        log.debug("file_parser.text_parsed", path=file_path, encoding=encoding)
        return text


# ── Helpers ───────────────────────────────────────────────────────────────────

def _flatten_markdown_tables(text: str) -> str:
    """
    Xử lý markdown tables để chunker không cắt ngang bảng.

    Vấn đề: RecursiveCharacterTextSplitter dùng \\n làm separator
    → header row và data rows thành chunk riêng → mất context bảng.

    Giải pháp: Gom toàn bộ bảng thành 1 khối văn bản liên tục,
    mỗi data row được prefix bằng header để tự đủ nghĩa:
      "Policy: HR-001 | Mô tả: Nghỉ phép năm | Ngày hiệu lực: 2024-01-01"

    Điều này cho phép mỗi data row được tìm kiếm độc lập kể cả khi bị
    tách vào chunk khác nhau.
    """
    import re

    lines = text.splitlines()
    result: list[str] = []
    i = 0

    while i < len(lines):
        line = lines[i]

        # Phát hiện dòng đầu của bảng markdown (có ít nhất 2 ký tự |)
        if re.match(r"^\s*\|.+\|", line):
            table_lines: list[str] = []

            # Thu thập toàn bộ dòng liên tiếp thuộc bảng
            while i < len(lines) and re.match(r"^\s*\|", lines[i]):
                table_lines.append(lines[i])
                i += 1

            if len(table_lines) < 2:
                # Không đủ dòng để xử lý → giữ nguyên
                result.extend(table_lines)
                continue

            # Dòng 0: header  |  Dòng 1: separator (|---|---|)  |  Dòng 2+: data
            header_row = table_lines[0]
            separator_idx = 1 if re.match(r"^\s*\|[-:\s|]+\|", table_lines[1]) else None
            data_start = 2 if separator_idx is not None else 1

            # Parse cột header
            headers = [
                h.strip()
                for h in header_row.strip().strip("|").split("|")
                if h.strip()
            ]

            flat_rows: list[str] = []

            for data_line in table_lines[data_start:]:
                cells = [
                    c.strip()
                    for c in data_line.strip().strip("|").split("|")
                ]
                # Pair từng cell với header tương ứng
                if headers:
                    pairs = [
                        f"{headers[j]}: {cells[j]}"
                        for j in range(min(len(headers), len(cells)))
                        if cells[j]
                    ]
                    if pairs:
                        flat_rows.append(" | ".join(pairs))
                elif any(cells):
                    flat_rows.append(" | ".join(c for c in cells if c))

            if flat_rows:
                # Giữ header gốc (để con người đọc dễ) + các dòng đã flatten
                result.append(header_row.strip())
                result.extend(flat_rows)
                result.append("")  # blank line để tách với block tiếp theo
            else:
                # Bảng trống → giữ nguyên
                result.extend(table_lines)

        else:
            result.append(line)
            i += 1

    return "\n".join(result)


def _ext_from_mime(mime_type: str) -> Optional[str]:
    # .doc (application/msword) intentionally omitted — use .docx only
    MIME_MAP = {
        "application/pdf": "pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "docx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "xlsx",
        "application/vnd.ms-excel": "xls",
        "text/plain": "txt",
        "text/markdown": "md",
    }
    return MIME_MAP.get(mime_type)


def _ext_from_name(file_name: str) -> Optional[str]:
    if "." not in file_name:
        return None
    return file_name.rsplit(".", 1)[-1].lower()
